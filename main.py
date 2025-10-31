import sys
from PySide6.QtWidgets import (
    QApplication, QDialog, QFormLayout, QLineEdit, QTextEdit, QDateEdit,
    QComboBox, QPushButton, QHBoxLayout, QVBoxLayout, QMessageBox, QFileDialog, QWidget, QLabel
)
from PySide6.QtCore import Qt, QDate, Slot, QRegularExpression
from PySide6.QtGui import QRegularExpressionValidator
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class ClienteDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestor de Clientes ART")
        self.setMinimumWidth(720)
        self.setMinimumHeight(520)

        # Paletas de color (puedes ajustar a gusto)
        self.themes = {
            "Rojo profesional": {
                "bg": "#f4f4f6",
                "card": "#ffffff",
                "primary": "#C70000",
                "primary_hover": "#b10000",
                "primary_pressed": "#990000",
                "text": "#1f2937",
                "muted": "#6b7280",
                "border": "#e5e7eb",
                "focus": "#ef4444",
                "header_bg": "#ffecec"
            },
            "Colore Secundario": {
                "bg": "#f6f5ff",
                "card": "#ffffff",
                "primary": "rgb(147, 51, 234)",            # morado
                "primary_hover": "rgb(132, 40, 228)",
                "primary_pressed": "rgb(116, 34, 201)",
                "text": "#111827",
                "muted": "#6b7280",
                "border": "#e5e7eb",
                "focus": "rgb(147, 51, 234)",
                "header_bg": "#f0e9ff"
            }
        }

        self.init_ui()
        self.apply_theme("Rojo profesional")  # tema inicial

    def init_ui(self):
        # ====== HEADER ======
        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(16, 12, 16, 12)
        self.lblTitle = QLabel("GESTOR DE CLIENTES ART")
        self.lblTitle.setObjectName("Title")
        header_layout.addWidget(self.lblTitle)
        header_layout.addStretch(1)

        self.cboTheme = QComboBox()
        self.cboTheme.addItems(self.themes.keys())
        self.cboTheme.currentTextChanged.connect(self.apply_theme)
        header_layout.addWidget(self.cboTheme)

        # ====== CARD (contenedor blanco) ======
        self.card = QWidget()
        card_layout = QVBoxLayout(self.card)
        card_layout.setContentsMargins(20, 20, 20, 20)
        card_layout.setSpacing(16)

        # ----- Campos -----
        self.txtNombre = QLineEdit(placeholderText="Ej: Juan")
        self.txtNombre.setMaxLength(60)

        self.txtApellido = QLineEdit(placeholderText="Ej: Pérez")
        self.txtApellido.setMaxLength(60)

        self.txtDni = QLineEdit(placeholderText="Solo números")
        self.txtDni.setMaxLength(15)
        dni_regex = QRegularExpression(r"^\d{0,15}$")  # permite escribir hasta 15 dígitos
        self.txtDni.setValidator(QRegularExpressionValidator(dni_regex))

        self.dateAlta = QDateEdit()
        self.dateAlta.setCalendarPopup(True)
        self.dateAlta.setDisplayFormat("yyyy-MM-dd")
        self.dateAlta.setDate(QDate.currentDate())

        self.txtLesion = QLineEdit(placeholderText="Ej: Esguince de tobillo")
        self.txtLesion.setMaxLength(80)

        self.txtEstudios = QTextEdit()
        self.txtEstudios.setPlaceholderText("Descripción breve de estudios médicos realizados...")

        self.cboArt = QComboBox()
        self.cboArt.setEditable(True)
        self.cboArt.addItems([
            "", "Prevención ART", "Swiss Medical ART", "Galeno ART", "Provincia ART",
            "Experta ART", "Mapfre ART", "La Segunda ART"
        ])
        self.cboArt.setCurrentIndex(0)

        # ----- Layout de formulario -----
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignLeft)
        form.setFormAlignment(Qt.AlignTop)
        form.setHorizontalSpacing(14)
        form.setVerticalSpacing(10)
        form.addRow("Nombre:", self.txtNombre)
        form.addRow("Apellido:", self.txtApellido)
        form.addRow("DNI:", self.txtDni)
        form.addRow("Día de Alta Médica:", self.dateAlta)
        form.addRow("Lesión:", self.txtLesion)
        form.addRow("Estudios Médicos:", self.txtEstudios)
        form.addRow("ART del cliente:", self.cboArt)

        card_layout.addLayout(form)

        # ----- Botones -----
        self.btnAgregar = QPushButton("Agregar")
        self.btnAgregar.setObjectName("PrimaryButton")
        self.btnAgregar.clicked.connect(self.on_agregar)

        self.btnLimpiar = QPushButton("Limpiar")
        self.btnLimpiar.clicked.connect(self.on_limpiar)

        btns = QHBoxLayout()
        btns.addStretch(1)
        btns.addWidget(self.btnLimpiar)
        btns.addWidget(self.btnAgregar)
        card_layout.addLayout(btns)

        # ====== ROOT ======
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)
        root.addWidget(header)
        root.addWidget(self.card)

    def apply_theme(self, name: str):
        t = self.themes.get(name, next(iter(self.themes.values())))
        css = f"""
        QDialog {{
            background: {t['bg']};
        }}
        QWidget#TitleBar {{ background: {t['header_bg']}; }}

        QLabel#Title {{
            font-size: 18px;
            font-weight: 700;
            letter-spacing: 0.5px;
            color: {t['text']};
        }}

        /* Card */
        QWidget {{
            font-size: 14px;
            color: {t['text']};
        }}
        QWidget#{self.card.objectName()} {{
            background: {t['card']};
            border: 1px solid {t['border']};
            border-radius: 14px;
        }}

        /* Inputs */
        QLineEdit, QComboBox, QDateEdit, QTextEdit {{
            background: #ffffff;
            border: 1px solid {t['border']};
            border-radius: 8px;
            padding: 8px 10px;
        }}
        QLineEdit:hover, QComboBox:hover, QDateEdit:hover, QTextEdit:hover {{
            border-color: #d1d5db;
        }}
        QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QTextEdit:focus {{
            border: 2px solid {t['focus']};
            outline: none;
        }}
        QComboBox::drop-down {{
            border: none;
            width: 26px;
        }}

        /* Labels del Form */
        QFormLayout > QLabel {{
            color: {t['muted']};
            font-weight: 600;
        }}

        /* Botón primario */
        QPushButton#PrimaryButton {{
            background: {t['primary']};
            color: white;
            border: none;
            border-radius: 10px;
            padding: 8px 16px;
            font-weight: 600;
        }}
        QPushButton#PrimaryButton:hover {{
            background: {t['primary_hover']};
        }}
        QPushButton#PrimaryButton:pressed {{
            background: {t['primary_pressed']};
        }}

        /* Botón secundario */
        QPushButton {{
            padding: 8px 16px;
            border-radius: 10px;
            background: #ffffff;
            border: 1px solid {t['border']};
            color: {t['text']};
        }}
        QPushButton:hover {{
            background: #fafafa;
        }}
        """
        self.setStyleSheet(css)
        # header color
        self.findChild(QWidget).setStyleSheet(f"background:{t['header_bg']};")

    # ----------------- Lógica original -----------------
    def _leer_campos(self):
        data = {
            "Nombre": self.txtNombre.text().strip(),
            "Apellido": self.txtApellido.text().strip(),
            "DNI": self.txtDni.text().strip(),
            "DiaAltaMedica": self.dateAlta.date().toString("yyyy-MM-dd"),
            "Lesion": self.txtLesion.text().strip(),
            "EstudiosMedicos": self.txtEstudios.toPlainText().strip(),
            "ART": (self.cboArt.currentText() or "").strip()
        }
        return data

    def _validar(self, data):
        faltan = []
        for k in ("Nombre", "Apellido", "DNI", "DiaAltaMedica", "ART"):
            if not data[k]:
                faltan.append(k)

        if data["DNI"] and not data["DNI"].isdigit():
            return False, "El DNI debe contener solo números."

        if faltan:
            return False, "Faltan datos obligatorios: " + ", ".join(faltan)
        return True, ""

    @Slot()
    def on_agregar(self):
        data = self._leer_campos()
        ok, msg = self._validar(data)
        if not ok:
            QMessageBox.warning(self, "Validación", msg)
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            caption="Guardar/Agregar en Excel",
            dir="clientes_art.xlsx",
            filter="Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            self._guardar_excel(path, data)
        except Exception as ex:
            QMessageBox.critical(self, "Error al guardar", f"No se pudo guardar el Excel.\n\n{ex}")
            return

        QMessageBox.information(self, "OK", "Datos guardados correctamente en el Excel.")

    def _guardar_excel(self, path, data):
        headers = [
            "Nombre", "Apellido", "DNI", "DiaAltaMedica",
            "Lesion", "EstudiosMedicos", "ART"
        ]

        try:
            wb = load_workbook(path)
            ws = wb.active
            if ws.max_row == 1 and all((ws.cell(row=1, column=i+1).value is None) for i in range(len(headers))):
                ws.append(headers)
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"
            ws.append(headers)

        row = [data[h] for h in headers]
        ws.append(row)

        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(16, len(h) + 2)

        wb.save(path)

    @Slot()
    def on_limpiar(self):
        self.txtNombre.clear()
        self.txtApellido.clear()
        self.txtDni.clear()
        self.txtLesion.clear()
        self.txtEstudios.clear()
        self.cboArt.setCurrentIndex(0)
        self.dateAlta.setDate(QDate.currentDate())


if __name__ == "__main__":
    app = QApplication(sys.argv)
    dlg = ClienteDialog()
    dlg.show()
    sys.exit(app.exec())
